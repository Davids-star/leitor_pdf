import os
import pytesseract
from pdf2image import convert_from_path
from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook

class LeitorDocumento:
    """Classe responsável por ler diferentes formatos de arquivos e extrair texto."""

    def __init__(self, poppler_path=None, tesseract_path=None):
        self.poppler_path = poppler_path
        if tesseract_path:
            pytesseract.pytesseract.tesseract_cmd = tesseract_path

    def ler_pdf(self, caminho):
        texto = ""
        try:
            # Limite de 10MB para evitar estouro de memória
            tamanho = os.path.getsize(caminho)
            if tamanho > 10 * 1024 * 1024:
                print(f"  ⚠ Arquivo muito grande para OCR ({tamanho/1024/1024:.1f}MB): {os.path.basename(caminho)}")
                return ""

            # Tenta extração direta de texto
            reader = PdfReader(caminho)
            for pagina in reader.pages:
                texto += pagina.extract_text() or ""

            # Se o texto estiver vazio, tenta OCR
            if not texto.strip():
                print(f"  [OCR] Ativado para: {os.path.basename(caminho)}")
                imagens = convert_from_path(caminho, dpi=100, first_page=1, last_page=1, poppler_path=self.poppler_path)
                for img in imagens:
                    texto += pytesseract.image_to_string(img)
        except Exception as e:
            print(f"  ❌ Erro ao ler PDF {os.path.basename(caminho)}: {e}")
        return texto

    def ler_txt(self, caminho):
        try:
            with open(caminho, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
        except Exception as e:
            print(f"  ❌ Erro ao ler TXT {os.path.basename(caminho)}: {e}")
            return ""

    def ler_docx(self, caminho):
        texto = ""
        try:
            if os.path.getsize(caminho) > 10 * 1024 * 1024:
                print(f"  ⚠ DOCX muito grande.")
                return ""
            doc = Document(caminho)
            for paragrafo in doc.paragraphs:
                texto += paragrafo.text + "\n"
        except Exception as e:
            print(f"  ❌ Erro ao ler DOCX {os.path.basename(caminho)}: {e}")
        return texto

    def ler_excel(self, caminho):
        texto = ""
        try:
            if os.path.getsize(caminho) > 10 * 1024 * 1024:
                print(f"  ⚠ Excel muito grande.")
                return ""
            if caminho.lower().endswith(".xls"):
                return "[ERRO] Formato .xls não suportado. Converta para .xlsx."
            
            wb = load_workbook(caminho, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for linha in ws.iter_rows(values_only=True):
                    for celula in linha:
                        if celula is not None:
                            texto += str(celula) + " "
        except Exception as e:
            print(f"  ❌ Erro ao ler Excel {os.path.basename(caminho)}: {e}")
        return texto
