#importações feitas para a automatização funcionar
import os
import re
import pytesseract
import difflib
import unicodedata
from docx import Document
from pdf2image import convert_from_path
from PyPDF2 import PdfReader
from openpyxl import load_workbook  
from buscar_pc import procurar_arquivos
from config import carregar_configuracoes, salvar_configuracoes, configurar_caminhos_manualmente

# Aque pega a pasta em que está todo o projeto
BASE = os.path.dirname(__file__)

# Configuração do Tesseract e Poppler
pytesseract.pytesseract.tesseract_cmd = os.path.join(
    BASE, "tesseract", "tesseract.exe"
)

POPPLER_PATH = os.path.join(BASE, "poppler","Library", "bin")

#  Minha classe para começar a ler os arquivos
class LeitorDocumento:

    #Responsavel por ler os arquivos e transformalos em texto
    # Arquivo para iniciar o programa e inditificando o poppler (ajuda o ocr a lê imagens) está

    def __init__(self, poppler_path=r"C:\Program Files\poppler\Library\bin"):
        self.poppler_path = poppler_path

    def ler_pdf(self, caminho):
        texto = ""
        try:

            # Ao ler com OCR, ele verifica tamanho do arquivo para evitar estouro de memória (Limite: 10MB)

            tamanho = os.path.getsize(caminho)
            if tamanho > 10 * 1024 * 1024:
                print(f"  ⚠ Arquivo muito grande para OCR/Processamento ({tamanho/1024/1024:.1f}MB): {os.path.basename(caminho)}")
                return ""

            #Soma cada pagina lida para acumular o texto e fazer a procura da palavra

            reader = PdfReader(caminho)
            for pagina in reader.pages:
                texto += pagina.extract_text() or ""

            # Se o texto estiver vazio, tenta OCR

            if not texto.strip():

                print(f"  [OCR] Ativado para: {os.path.basename(caminho)}")
                imagens = convert_from_path(caminho, dpi=100, first_page=1, last_page=1, poppler_path=self.poppler_path)
                for img in imagens:
                    texto += pytesseract.image_to_string(img) # A mesma logica, soma as imagens para transforma-lós em texto e lê, caso se não passar de 10MB.
        except Exception as e:
            print(f"Erro ao ler PDF {caminho}: {e}")
        return texto

    #funcões para lê txt, docx, xlsx ou xls

    def ler_txt(self, caminho):
        try:
            with open(caminho, 'r', encoding='utf-8') as f:
                return f.read()
        except Exception as e:
            print(f"Erro ao ler TXT {caminho}: {e}")
            return ""

    def ler_docx(self, caminho):
        texto = ""
        try:
            # Verifica tamanho (10MB)
            if os.path.getsize(caminho) > 10 * 1024 * 1024:
                print(f"  ⚠ DOCX ignorado por ser muito grande.")
                return ""

            doc = Document(caminho)
            for paragrafo in doc.paragraphs:
                texto += paragrafo.text + "\n"
        except Exception as e:
            print(f"Erro ao ler DOCX {caminho}: {e}")
        return texto

    def ler_excel(self, caminho):
        texto = ""
        try:
            # Verifica tamanho (10MB)
            if os.path.getsize(caminho) > 10 * 1024 * 1024:
                print(f"  ⚠ Excel ignorado por ser muito grande.")
                return ""

            # Verifica se é .xls (formato antigo)
            if caminho.lower().endswith(".xls"):
                return "[ERRO] Formato .xls não suportado. Converta para .xlsx ou instale 'xlrd'."

            wb = load_workbook(caminho, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for linha in ws.iter_rows(values_only=True):
                    for celula in linha:
                        if celula is not None:
                            texto += str(celula) + " "
            return texto
        except Exception as e:
            print(f"Erro ao ler Excel {caminho}: {e}")
            return ""

class BuscadorTexto:

    #Responsável por processar o texto e encontrar padrões.
    #Deixar todas as letras minusculas, tirar acentos e caracteres especiais
    def normalizar(self, texto):
        return unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("utf-8").lower() 

    def contem_termo(self, texto, termo):
        texto_norm = self.normalizar(texto)
        termo_norm = self.normalizar(termo)
        
        # Busca exata do termo completo

        if termo_norm in texto_norm:
            return True
        
        # Busca todas as palavras do termo separadamente, e palavras com até 80% de similaridade
        # Busca por palavras e similaridade

        palavras_texto = texto_norm.split()
        palavras_busca = termo_norm.split()
        
        for p_busca in palavras_busca:
            encontrou_similar = False
            for p_texto in palavras_texto:
                similaridade = difflib.SequenceMatcher(None, p_busca, p_texto).ratio()
                if similaridade > 0.8: #80%
                    encontrou_similar = True
                    break
            if not encontrou_similar:
                return False # Se uma das palavras da busca não foi achada (nem similar), falha
        
        return True

    #Funções para buscar padrões específicos

    def buscar_cpf(self, texto):
        return re.findall(r"\d{3}\.\d{3}\.\d{3}-\d{2}", texto)

    def buscar_telefone(self, texto):
        return re.findall(r"(?:\(?\d{2}\)?\s?)?(?:9\s?)?\d{4}[-\s]?\d{4}", texto)

    def buscar_stmb(self, texto):
        return re.findall(r"STMB[\s-]?\d+", texto, re.IGNORECASE)

    def buscar_numeros(self, texto):
        return re.findall(r"\d+", texto)

    def extrair_numero(self, texto):
        return re.findall(r"\d+", texto)




class ProcessadorArquivos:

    #Orquestrador que gerencia a pasta de entrada e utiliza o Leitor e o Buscador.

    def __init__(self, entrada=None):
        # Aqui faz a busca
        self.entrada = entrada
        self.leitor = LeitorDocumento()
        self.buscador = BuscadorTexto()

    def configurar_busca_pc(self):
        """Ativa a busca em pastas configuradas (Documents, Downloads, etc)."""
        config = carregar_configuracoes()
        print(f"🔍 Iniciando varredura no PC em: {', '.join(config['caminhos_busca'])}")
        
        todos_arquivos = []
        for caminho_base in config["caminhos_busca"]:
            # Usamos a procurar_arquivos de buscar_pc.py
            # Note que a procurar_arquivos original recebia base_path e pastas separadas.
            # Vamos adaptar para passar o caminho completo como base e "." como pasta.
            arquivos = procurar_arquivos(
                caminho_base, 
                ["."], 
                config["extensoes"]
            )
            todos_arquivos.extend(arquivos)
        
        self.entrada = todos_arquivos

    def processar_tudo(self, termo="", tipo_busca="texto"):
        encontrados = {}
        
        # Se a entrada for None, fazemos a varredura no PC agora
        if self.entrada is None:
            self.configurar_busca_pc()

        # Define a lista de caminhos a processar
        if isinstance(self.entrada, str):
            if not os.path.exists(self.entrada):
                print(f"Pasta {self.entrada} não encontrada.")
                return {}
            caminhos_completos = [os.path.join(self.entrada, f) for f in os.listdir(self.entrada)]
        else:
            caminhos_completos = self.entrada

        for caminho in caminhos_completos:
            if not os.path.isfile(caminho):
                continue
                
            arquivo_nome = os.path.basename(caminho)
            texto = ""

            # 1. Identifica e lê o arquivo(txt, docx, pdf, xlsx, xls)
            parent_dir = os.path.basename(os.path.dirname(caminho))
            print(f"Processando: [{parent_dir}] {arquivo_nome}")
            if arquivo_nome.lower().endswith(".pdf"):
                texto = self.leitor.ler_pdf(caminho)
            elif arquivo_nome.lower().endswith(".txt"):
                texto = self.leitor.ler_txt(caminho)
            elif arquivo_nome.lower().endswith(".docx"):
                texto = self.leitor.ler_docx(caminho)
            elif arquivo_nome.lower().endswith((".xlsx", ".xls")):
                texto = self.leitor.ler_excel(caminho)
            else:
                continue
            
            # 2. Realiza a busca baseada no tipo solicitado  

            res = [] 
            sucesso = False
            
            if tipo_busca == "cpf":
                res = self.buscador.buscar_cpf(texto)
                if res:
                    print(f"  ✔ CPF(s): {res}")
                    sucesso = True
            elif tipo_busca == "telefone":
                res = self.buscador.buscar_telefone(texto)
                if res:
                    print(f"  ✔ Telefone(s): {res}")
                    sucesso = True
            elif tipo_busca == "numero":
                res = self.buscador.buscar_numeros(texto)
                if res:
                    print(f"  ✔ Números: {res[:5]}...")
                    sucesso = True
            elif tipo_busca == "stmb":
                res = self.buscador.buscar_stmb(texto)
                if res:
                    print(f"  ✔ STMB: {res}")
                    sucesso = True
            elif tipo_busca == "extrair_numero":
                res = self.buscador.extrair_numero(texto)
                if res:
                    print(f"  ✔ Número: {res[:10]}")
                    sucesso = True
            else:  # tipo_busca == "texto"
                if self.buscador.contem_termo(texto, termo):
                    print(f"  ✔ Termo encontrado!")
                    res = [termo]
                    sucesso = True
                else:
                    print(f"  ✘ Termo não encontrado.")
                

            if sucesso:
                encontrados[caminho] = res

        return encontrados

    def buscar_em_lista_especifica(self, lista_nomes, termo):

        # Garante que temos uma lista de arquivos

        if self.entrada is None:
            self.configurar_busca_pc()
            
        if isinstance(self.entrada, str):
            arquivos_disponiveis = os.listdir(self.entrada)
            base = self.entrada
        else:
            arquivos_disponiveis = self.entrada
            base = ""

        arquivo_filtrados = []
        for arquivo in arquivos_disponiveis:
            nome_arquivo = os.path.basename(arquivo)
            if any(n.lower() in nome_arquivo.lower() for n in lista_nomes):
                arquivo_filtrados.append(arquivo)

        encontrados = {}
        for caminho in arquivo_filtrados:
            caminho_completo = os.path.join(base, caminho) if base else caminho
            
            if not os.path.exists(caminho_completo):
                print(f"Arquivo não encontrado: {caminho}")
                continue
                
            print(f"\nBuscando termo em: {os.path.basename(caminho_completo)}")
            # Tenta ler de acordo com a extensão
            ext = caminho_completo.lower()
            if ext.endswith(".pdf"):
                texto = self.leitor.ler_pdf(caminho_completo)
            elif ext.endswith(".txt"):
                texto = self.leitor.ler_txt(caminho_completo)
            elif ext.endswith(".docx"):
                texto = self.leitor.ler_docx(caminho_completo)
            elif ext.endswith((".xlsx", ".xls")):
                texto = self.leitor.ler_excel(caminho_completo)
            else:
                texto = "" 

            if self.buscador.contem_termo(texto, termo):
                print("  ✔ Encontrado!")
                encontrados[caminho_completo] = [termo]
            else:
                print("  ✘ Não encontrado.")
        return encontrados
    # Aqui onde fazemos a chamada de todas as funções que fizemos
def main():
    app = ProcessadorArquivos()

    print("\n" + "="*40)
    print("      BUSCADOR GLOBAL DE DOCUMENTOS")
    print("="*40)
    
    termo = input("\nDigite o que deseja encontrar no seu PC: ")

    print("\nO QUE DESEJA BUSCAR?")
    print("1 - O termo digitado (Busca de texto completo)")
    print("2 - Buscar CPFs")
    print("3 - Buscar Telefones")
    print("4 - Buscar em arquivos específicos (Filtrar por nome)")
    print("5 - Buscar STMBs")
    print("6 - Extrair todos os números")
    print("7 - Configurar pastas de busca")

    opcao = input("\nEscolha uma opção: ")

    if opcao == "1":
        resultados = app.processar_tudo(termo)
        print("\nArquivos encontrados com o termo:")
        for i, (caminho, achados) in enumerate(resultados.items()):
            nome = os.path.basename(caminho)
            pasta = os.path.dirname(caminho)
            print(f"{i} - {nome} (Pasta: {pasta})")

    elif opcao == "2":
        resultados = app.processar_tudo("", tipo_busca="cpf")
        print("\nCPFs Encontrados:")
        for caminho, cpfs in resultados.items():
            nome = os.path.basename(caminho)
            pasta = os.path.dirname(caminho)
            print(f"\n📄 Arquivo: {nome}")
            print(f"📍 Local: {pasta}")
            for cpf in cpfs:
                print(f"  ✔ {cpf}")

    elif opcao == "3":
        resultados = app.processar_tudo("", tipo_busca="telefone")
        print("\nTelefones Encontrados:")
        for caminho, telefones in resultados.items():
            nome = os.path.basename(caminho)
            pasta = os.path.dirname(caminho)
            print(f"\n📄 Arquivo: {nome}")
            print(f"📍 Local: {pasta}")
            for tel in telefones:
                print(f"  ✔ {tel}")

    elif opcao == "4":
        nomes = input("Digite parte do nome dos arquivos (separados por vírgula): ").split(",")
        nomes = [n.strip() for n in nomes]
        resultados = app.buscar_em_lista_especifica(nomes, termo)
        print("\nArquivos filtrados encontrados com o termo:")
        for caminho, achados in resultados.items():
            nome = os.path.basename(caminho)
            pasta = os.path.dirname(caminho)
            print(f"📄 Arquivo: {nome}")
            print(f"📍 Local: {pasta}")
            

    elif opcao == "5":
        resultados = app.processar_tudo("", tipo_busca="stmb")
        print("\nSTMBs Encontrados:")
        for caminho, stmbs in resultados.items():
            nome = os.path.basename(caminho)
            pasta = os.path.dirname(caminho)
            print(f"\n📄 Arquivo: {nome}")
            print(f"📍 Local: {pasta}")
            for stmb in stmbs:
                print(f"  ✔ {stmb}")

    elif opcao == "6":
        resultados = app.processar_tudo("", tipo_busca="extrair_numero")
        print("\nExtração de Números Concluída:")
        for caminho, numeros in resultados.items():
            nome = os.path.basename(caminho)
            pasta = os.path.dirname(caminho)
            print(f" - {nome} ({len(numeros)} números encontrados) -> {pasta}")

    elif opcao == "7":
        configurar_caminhos_manualmente()

    print("\n" + "="*40)
    input("Pressione ENTER para fechar o programa...")

if __name__ == "__main__":
    main()